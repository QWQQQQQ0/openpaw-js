"""Excel document generator using openpyxl."""

from __future__ import annotations

from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


@dataclass
class CellData:
    value: str | int | float | None = None
    bold: bool = False
    header: bool = False


@dataclass
class SheetData:
    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list[str | int | float | None]] = field(default_factory=list)


class ExcelGenerator:
    """Generate Excel spreadsheets from structured data."""

    def generate(
        self,
        title: str,
        sheets: list[dict],
        author: str | None = None,
    ) -> bytes:
        """Generate an Excel workbook.

        Args:
            title: Workbook title (used as first sheet name if no sheets provided)
            sheets: List of {name: str, headers: [str], rows: [[values]]}
            author: Optional author metadata

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        # Header style
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Border style
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for sheet_data in sheets:
            name = sheet_data.get("name", "Sheet")
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            ws = wb.create_sheet(title=name)

            # Add headers
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Add data rows
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
            else:
                # No headers, just add rows
                for row_idx, row in enumerate(rows, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[column].width = adjusted_width

        # If no sheets provided, create a default one
        if not sheets:
            ws = wb.create_sheet(title=title)
            ws.cell(row=1, column=1, value="No data provided")

        # Serialize to bytes
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def generate_from_csv_text(
        self,
        title: str,
        csv_content: str,
        sheet_name: str = "Sheet1",
    ) -> bytes:
        """Generate Excel from CSV-like text.

        Args:
            title: Workbook title
            csv_content: CSV text (comma-separated)
            sheet_name: Name for the sheet

        Returns:
            Excel file as bytes
        """
        lines = csv_content.strip().split("\n")
        if not lines:
            return self.generate(title, [])

        headers = [h.strip() for h in lines[0].split(",")]
        rows = []
        for line in lines[1:]:
            row = []
            for val in line.split(","):
                val = val.strip()
                # Try to convert to number
                try:
                    row.append(int(val))
                except ValueError:
                    try:
                        row.append(float(val))
                    except ValueError:
                        row.append(val)
            rows.append(row)

        return self.generate(title, [{
            "name": sheet_name,
            "headers": headers,
            "rows": rows,
        }])


# Standalone usage
if __name__ == "__main__":
    gen = ExcelGenerator()
    sheets = [
        {
            "name": "销售数据",
            "headers": ["产品", "Q1", "Q2", "Q3", "Q4", "总计"],
            "rows": [
                ["产品A", 100, 120, 130, 150, 500],
                ["产品B", 80, 90, 100, 110, 380],
                ["产品C", 200, 180, 220, 250, 850],
            ],
        },
        {
            "name": "月度统计",
            "headers": ["月份", "收入", "支出", "利润"],
            "rows": [
                ["1月", 10000, 7000, 3000],
                ["2月", 12000, 8000, 4000],
                ["3月", 15000, 9000, 6000],
            ],
        },
    ]
    xlsx_bytes = gen.generate("年度销售报表", sheets)
    with open("test_output.xlsx", "wb") as f:
        f.write(xlsx_bytes)
    print("Generated test_output.xlsx")
